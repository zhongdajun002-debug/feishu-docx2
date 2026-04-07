# !/usr/bin/env python
# -*- coding: utf-8 -*-
# =====================================================
# @File   ：sheet.py
# @Date   ：2025/01/09 18:30
# @Author ：leemysw
# 2025/01/09 18:30   Create
# 2026/04/07 10:00   Add export_to_excel() / export_to_csv() for binary export
# =====================================================
"""
[INPUT]: 依赖 feishu_docx2.core.sdk 的 FeishuSDK
[OUTPUT]: 对外提供 SheetParser 类，将飞书电子表格解析为 Markdown/Excel/CSV
[POS]: parsers 模块的电子表格解析器
[PROTOCOL]: 变更时更新此头部，然后检查 CLAUDE.md
"""

import csv
from pathlib import Path
from typing import List, Optional

from feishu_docx2.core.sdk import FeishuSDK
from feishu_docx2.schema.models import TableMode
from feishu_docx2.utils.progress import ProgressManager
from feishu_docx2.utils.render_table import extract_cell_value


class SheetParser:
    """
    飞书电子表格解析器

    将飞书电子表格解析为 Markdown 格式，每个工作表作为一个章节。
    """

    def __init__(
            self,
            spreadsheet_token: str,
            user_access_token: str,
            table_mode: str = "md",
            sdk: Optional[FeishuSDK] = None,
            silent: bool = False,
            progress_callback=None,
    ):
        """
        初始化电子表格解析器

        Args:
            spreadsheet_token: 电子表格 token
            user_access_token: 用户访问凭证
            table_mode: 表格输出格式 ("html" 或 "md")
            sdk: 可选的 SDK 实例
            silent: 是否静默模式
            progress_callback: 进度回调函数
        """
        self.sdk = sdk or FeishuSDK()
        self.table_mode = TableMode(table_mode)
        self.user_access_token = user_access_token
        self.spreadsheet_token = spreadsheet_token
        self.block_info = {}

        # 进度管理器
        self.pm = ProgressManager(silent=silent, callback=progress_callback)

    def parse(self) -> str:
        """
        解析电子表格为 Markdown

        Returns:
            Markdown 格式的内容，每个工作表作为一个章节
        """
        pm = self.pm

        # 获取工作表列表
        with pm.spinner("获取工作表列表..."):
            sheets = self.sdk.sheet.get_sheet_list(
                spreadsheet_token=self.spreadsheet_token,
                access_token=self.user_access_token,
            )

        total_sheets = len(sheets)
        pm.log(f"  [dim]发现 {total_sheets} 个工作表[/dim]")
        pm.report("发现工作表", total_sheets, total_sheets)

        if total_sheets == 0:
            return ""

        sections = []

        # 解析每个工作表
        with pm.bar("解析工作表...", total_sheets) as advance:
            for sheet in sheets:
                sheet_id = sheet.sheet_id
                sheet_title = sheet.title
                resource_type = sheet.resource_type

                sheet_data = None

                if resource_type == "sheet":
                    sheet_data = self.sdk.sheet.get_sheet(
                        sheet_token=self.spreadsheet_token,
                        sheet_id=sheet_id,
                        access_token=self.user_access_token,
                        table_mode=self.table_mode,
                    )
                elif resource_type == "bitable":
                    sheet_data = self._parse_bitable_sheet(sheet_id, sheet_title)
                else:
                    pm.log(f"  [yellow]跳过不支持类型: {resource_type}[/yellow]")

                if sheet_data:
                    sections.append(f"# {sheet_title}\n\n{sheet_data}")

                advance()  # noqa

        pm.log(f"  [dim]解析完成 ({len(sections)} 个工作表)[/dim]")
        pm.report("解析完成", len(sections), total_sheets)

        return "\n\n---\n\n".join(sections)

    def _parse_bitable_sheet(self, sheet_id: str, sheet_title: str) -> Optional[str]:
        """解析嵌入的 Bitable 工作表"""
        pm = self.pm

        # 获取 block info
        if not self.block_info:
            blocks = self.sdk.sheet.get_sheet_metadata(
                spreadsheet_token=self.spreadsheet_token,
                access_token=self.user_access_token,
            )
            if blocks:
                for block in blocks:
                    block_info = block.get("blockInfo")
                    if block_info:
                        block_token = block_info.get("blockToken", "")
                        self.block_info[block.get("sheetId")] = block_token

        token = self.block_info.get(sheet_id, "")
        if not token:
            pm.log(f"  [yellow]跳过: {sheet_title}[/yellow]")
            return None

        token_parts = token.split("_")
        if len(token_parts) < 2:
            pm.log(f"  [yellow]跳过无效 token: {sheet_title}[/yellow]")
            return None

        return self.sdk.bitable.get_bitable(
            app_token=token_parts[0],
            table_id=token_parts[1],
            access_token=self.user_access_token,
            table_mode=self.table_mode,
        )

    # ======================================================================
    # Excel / CSV 导出
    # ======================================================================

    def _get_all_sheet_data(self) -> List[tuple[str, list]]:
        """
        获取所有工作表的原始数据。

        Returns:
            list of (sheet_title, values) tuples
        """
        sheets = self.sdk.sheet.get_sheet_list(
            spreadsheet_token=self.spreadsheet_token,
            access_token=self.user_access_token,
        )
        result = []
        for sheet in sheets:
            if sheet.resource_type != "sheet":
                continue
            values = self.sdk.sheet.get_sheet_values(
                sheet_token=self.spreadsheet_token,
                sheet_id=sheet.sheet_id,
                access_token=self.user_access_token,
            )
            result.append((sheet.title or sheet.sheet_id, values or []))
        return result

    def export_to_excel(self, output_dir: Path, filename: str) -> Path:
        """
        将电子表格所有工作表导出为 Excel (.xlsx) 文件。

        Args:
            output_dir: 输出目录
            filename: 文件名（不含扩展名）

        Returns:
            输出文件路径
        """
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认空白 Sheet

        sheet_data = self._get_all_sheet_data()

        for title, values in sheet_data:
            # Excel Sheet 名称最多 31 个字符，且不能含特殊字符
            safe_title = title[:31]
            ws = wb.create_sheet(title=safe_title)
            for row in values:
                ws.append([extract_cell_value(cell) for cell in row])

        if not wb.sheetnames:
            wb.create_sheet("Sheet1")

        output_path = output_dir / f"{filename}.xlsx"
        output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def export_to_csv(self, output_dir: Path, filename: str) -> List[Path]:
        """
        将电子表格每个工作表导出为独立的 CSV 文件（UTF-8 with BOM，Excel 可直接打开）。

        单工作表时输出 {filename}.csv；
        多工作表时输出 {filename}_{sheet_title}.csv。

        Args:
            output_dir: 输出目录
            filename: 文件名前缀（不含扩展名）

        Returns:
            输出文件路径列表
        """
        sheet_data = self._get_all_sheet_data()
        output_dir.mkdir(parents=True, exist_ok=True)
        paths = []

        single = len(sheet_data) == 1

        for title, values in sheet_data:
            if single:
                csv_path = output_dir / f"{filename}.csv"
            else:
                safe_title = title.replace("/", "_").replace("\\", "_")[:50]
                csv_path = output_dir / f"{filename}_{safe_title}.csv"

            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                for row in values:
                    writer.writerow([extract_cell_value(cell) for cell in row])

            paths.append(csv_path)

        return paths
